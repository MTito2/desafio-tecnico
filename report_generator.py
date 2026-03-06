import os
import pandas as pd
from utils import save_log
from config import FILES_PATH
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_xlsx(content):
        
    try:    
        df = pd.DataFrame(content)

        new_names = {
            "country": "País",
            "population": "Qtd habitantes",
            "median_age": "Média idade",
            "capital": "Capital",
            "languages": "Linguagens",
            "currencies": "Moedas"
        }

        df = df.rename(columns=new_names)
        df = df.sort_values(by=['Média idade'], ascending=False)

        wb = Workbook()
        ws = wb.active
        ws.title = "Teste RPA"

        blue_onfly = "009efb"
        aligned_centered = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws.row_dimensions[1].height = 30
        ws.column_dimensions['A'].width = 16

        for col in range(1, 7):
            ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=blue_onfly)

        logo_path = FILES_PATH / "logo.png"
        img = Image(logo_path)
        img.width, img.height = 90, 50
        ws.add_image(img, "A1")

        bold_font = Font(color="000000", bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = aligned_centered
                cell.border = thin_border
                if r_idx == 2:
                    cell.font = bold_font

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            max_len_data = df[col_name].astype(str).map(len).max()
            adjusted_width = max(max_len, max_len_data, 12) + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        dir = os.path.join(FILES_PATH, "Teste RPA - Mateus da Cunha Tito.xlsx")
        wb.save(dir)

        save_log("Planilha gerada com sucesso", "info")
        print("Planilha gerada com sucesso")
    
    except Exception as error:
        save_log(error, "error")
        print("Erro encontrado, verifique o arquivo 'logs.txt' para mais detalhes")
